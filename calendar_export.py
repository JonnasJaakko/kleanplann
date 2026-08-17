"""Экспорт расписания и сводной отчётности в CSV/Excel."""
import csv
from project import Project
from cost_calculator import calculate_cost


def _room(project, task):
    if 0 <= task.floor_index < len(project.floors):
        return next((r for r in project.floors[task.floor_index].rooms if r.id == task.room_id), None)
    return None


def export_tasks_csv(project: Project, filename: str):
    cost = calculate_cost(project)
    with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, delimiter=';')
        writer.writerow(["KleanPlann — расписание"])
        writer.writerow(["Проект", project.name])
        writer.writerow(["Погода", {1.0:'Ясно',1.2:'Дождь',1.5:'Снег',1.8:'Сильный дождь'}.get(project.weather_factor, '—')])
        writer.writerow(["Тип уборки", getattr(project, 'cleaning_type', 'поддерживающая')])
        writer.writerow(["Надбавка за переработку, %", getattr(project, 'overtime_premium_percent', 50)])
        writer.writerow(["Стоимость", cost['cost_with_overtime']])
        writer.writerow([])
        writer.writerow(["Сотрудник", "Этаж", "Комната", "Тип", "Площадь (м²)", "Начало", "Окончание", "Продолжительность (мин)", "Статус"])
        for task in sorted(project.cleaning_tasks, key=lambda t: (t.employee, t.start_dt)):
            room = _room(project, task)
            duration = (task.end_dt-task.start_dt).total_seconds()/60
            writer.writerow([
                project.employee_names[task.employee] if task.employee < len(project.employee_names) else f"Сотрудник {task.employee+1}",
                task.floor_index+1,
                room.name if room else f"Комната {task.room_id+1}",
                room.room_type if room else "",
                f"{room.area_m2:.1f}" if room else "",
                task.start_dt.strftime('%Y-%m-%d %H:%M'), task.end_dt.strftime('%Y-%m-%d %H:%M'),
                f"{duration:.0f}", "СВЕРХ СМЕНЫ" if getattr(task,'is_overtime',False) else "В смене"
            ])


def export_tasks_excel(project: Project, filename: str):
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Расписание"
    cost = calculate_cost(project)
    ws.append(["KleanPlann — расписание"])
    ws.append(["Проект", project.name])
    ws.append(["Тип уборки", getattr(project, 'cleaning_type', 'поддерживающая')])
    ws.append(["Погода", {1.0:'Ясно',1.2:'Дождь',1.5:'Снег',1.8:'Сильный дождь'}.get(project.weather_factor, '—')])
    ws.append(["Надбавка за переработку, %", getattr(project, 'overtime_premium_percent', 50)])
    ws.append(["Стоимость", cost['cost_with_overtime']])
    ws.append([])
    headers=["Сотрудник","Этаж","Комната","Тип","Площадь (м²)","Начало","Окончание","Продолжительность (мин)","Статус"]
    ws.append(headers)
    for c in ws[8]: c.font=Font(bold=True)
    for task in sorted(project.cleaning_tasks, key=lambda t:(t.employee,t.start_dt)):
        room=_room(project,task); overtime=getattr(task,'is_overtime',False)
        ws.append([
            project.employee_names[task.employee] if task.employee < len(project.employee_names) else f"Сотрудник {task.employee+1}",
            task.floor_index+1, room.name if room else f"Комната {task.room_id+1}", room.room_type if room else "",
            room.area_m2 if room else None, task.start_dt.strftime('%Y-%m-%d %H:%M'), task.end_dt.strftime('%Y-%m-%d %H:%M'),
            (task.end_dt-task.start_dt).total_seconds()/60, "СВЕРХ СМЕНЫ" if overtime else "В смене"
        ])
        if overtime:
            for cell in ws[ws.max_row]: cell.fill=PatternFill('solid', fgColor='F4CCCC')
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width=min(32,max(12,max(len(str(c.value or '')) for c in col)+2))

    summary=wb.create_sheet("Сводка")
    summary.append(["Показатель","Значение"])
    summary['A1'].font=summary['B1'].font=Font(bold=True)
    summary_rows=[
        ("Общее время, ч",cost['total_time_hours']),
        ("Переработка, ч",cost['overtime_hours']),
        ("Стоимость, руб.",cost['cost_with_overtime']),
        ("Рекомендуемый штат",cost['needed_employees']),
        ("Сотрудников в проекте",project.employees_count),
        ("Погода",{1.0:'Ясно',1.2:'Дождь',1.5:'Снег',1.8:'Сильный дождь'}.get(project.weather_factor,'—')),
        ("Тип уборки",getattr(project,'cleaning_type','поддерживающая')),
    ]
    for row in summary_rows: summary.append(row)
    wb.save(filename)
